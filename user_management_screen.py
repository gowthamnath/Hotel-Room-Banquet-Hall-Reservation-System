import customtkinter as ctk
import mysql.connector
from mysql.connector import Error
from tkinter import messagebox, filedialog
from datetime import datetime
import csv
import os
import sys
import subprocess

# Professional Excel Logic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Database Configuration
DB_CONFIG = {
    "host": "141.209.241.57",
    "port": 3306,
    "user": "putch1v",
    "password": "mypass",
    "database": "BIS698MSpring26Group_2"
}

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("green")

# UI Styling Colors
PRIMARY_GREEN = "#5C8A3C"
SIDEBAR_GREEN = "#4A7A2E"
DARK_TEXT = "#2C3E50"
GRAY_TEXT = "#666666"
SECTION_GREEN = "#3D7A1A"

NAV_ITEMS = [
    ("User & Report Management", "👥"), ("Room & Hall Management", "🛏"),
    ("Booking Calendar", "📅"), ("BI Dashboard", "📊"), ("Settings", "⚙")
]

GUEST_CARDS = [
    ("➕", "Add Guest", "Create new guest accounts and profiles"),
    ("✏️", "Edit Guest", "Update guest information and preferences"),
    ("🗃", "View All Guests", "Browse and search guest database")
]

STAFF_CARDS = [
    ("👷", "Add Staff", "Create new staff accounts with roles"),
    ("🔑", "Manage Permissions", "Set access levels and permissions"),
    ("📊", "Staff Reports", "View staff activity and performance")
]

REPORT_CARDS = [
    ("💰", "Revenue & Payment Performance", "Track revenue and payment trends"),
    ("📊", "Occupancy & Utilization Report", "Room and banquet hall usage efficiency analysis"),
    ("👥", "Customer & Booking Behavior", "Guest activity, booking patterns, and popular customers")
]

# Standard Security Questions for Dropdowns
SECURITY_QUESTIONS = [
    "What was the name of your first pet?",
    "In what city were you born?",
    "What was the name of your elementary school?",
    "What is your mother's maiden name?",
    "What was the make of your first car?",
    "What is your favorite book?"
]
class UserManagementScreen(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.admin_id = self._get_session_id()
        self.admin_name = self._get_admin_name()
        self.title("Admin Center - User Management")
        
        # Set to Full Screen (Maximized) by default
        self.after(0, lambda: self.state('zoomed'))
        
        self.configure(fg_color="white")
        self._build_ui()

    def _get_session_id(self):
        path = os.path.join(os.path.dirname(__file__), "session.txt")
        return open(path, "r").read().strip() if os.path.exists(path) else None

    def _get_admin_name(self):
        if not self.admin_id: return "Admin"
        try:
            conn = mysql.connector.connect(**DB_CONFIG); cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT first_name, last_name FROM User WHERE user_id = %s", (self.admin_id,))
            res = cursor.fetchone(); conn.close()
            return f"{res['first_name']} {res['last_name']}".strip() if res else "Admin"
        except Error: return "Admin"

    def _build_ui(self):
        outer = ctk.CTkFrame(self, fg_color="white"); outer.pack(fill="both", expand=True)
        sidebar = ctk.CTkFrame(outer, fg_color=SIDEBAR_GREEN, width=215, corner_radius=0); sidebar.pack(side="left", fill="y"); sidebar.pack_propagate(False)
        ctk.CTkLabel(sidebar, text="⚙ Admin Center", font=ctk.CTkFont(size=15, weight="bold"), text_color="white").pack(pady=20, padx=15, anchor="w")
        
        # --- UPDATED SIDEBAR BUTTONS WITH NAVIGATION ---
        for label, icon in NAV_ITEMS:
            is_active = (label == "User Management")
            btn = ctk.CTkButton(sidebar, 
                                text=f"  {icon}  {label}", 
                                fg_color="#3D6A22" if is_active else "transparent", 
                                hover_color="#3D6A22", 
                                text_color="white", 
                                anchor="w", 
                                height=40,
                                command=lambda l=label: self._navigate(l))
            btn.pack(fill="x", padx=10, pady=2)
            
        ctk.CTkButton(sidebar, text="Sign Out", fg_color="transparent", border_width=1, border_color="#AAAAAA", command=self._sign_out).pack(side="bottom", padx=15, pady=20)

        main_scroll = ctk.CTkScrollableFrame(outer, fg_color="white"); main_scroll.pack(side="left", fill="both", expand=True)
        ctk.CTkLabel(main_scroll, text="User Management", font=ctk.CTkFont(family="Georgia", size=26, weight="bold"), text_color=SIDEBAR_GREEN).pack(anchor="w", padx=30, pady=(25, 5))
        
        welcome_frame = ctk.CTkFrame(main_scroll, fg_color="transparent"); welcome_frame.pack(fill="x", padx=30, pady=(0, 20))
        ctk.CTkLabel(welcome_frame, text=f"Welcome back, {self.admin_name}! 👋", font=ctk.CTkFont(size=16, weight="bold"), text_color=DARK_TEXT).pack(anchor="w")
        ctk.CTkLabel(welcome_frame, text="Centralized control for guests, staff, and Organizational Reports.", font=ctk.CTkFont(size=12), text_color=GRAY_TEXT).pack(anchor="w")

        self._make_section(main_scroll, "Guest Management", GUEST_CARDS)
        self._make_section(main_scroll, "Staff Management", STAFF_CARDS)
        self._make_section(main_scroll, "Reports", REPORT_CARDS)

    def _navigate(self, label):
        # Dictionary mapping labels to filenames
        file_map = {
            "User & Report Management": "user_management_screen.py",
            "Room & Hall Management": "room_management_admin.py",
            "Booking Calendar": "booking_calender.py",
            "BI Dashboard": "bi_dashboard.py",
            "Settings": "settings_admin.py"
        }
        
        # Don't do anything if we are already on this screen
        if label == "User Management":
            return
            
        target_file = file_map.get(label)
        if target_file:
            target_path = os.path.join(os.path.dirname(__file__), target_file)
            
            # Check if file exists before trying to open
            if os.path.exists(target_path):
                self.destroy()
                subprocess.Popen([sys.executable, target_path])
            else:
                messagebox.showerror("Navigation Error", f"File not found: {target_file}")

    def _make_section(self, parent, title, cards):
        section_frame = ctk.CTkFrame(parent, fg_color="white"); section_frame.pack(fill="x", padx=30, pady=(10, 0))
        ctk.CTkLabel(section_frame, text=title, font=ctk.CTkFont(size=17, weight="bold"), text_color=SECTION_GREEN).pack(anchor="w", pady=(0, 4))
        ctk.CTkFrame(section_frame, fg_color="#C8DDB4", height=1).pack(fill="x")
        cards_row = ctk.CTkFrame(parent, fg_color="white"); cards_row.pack(fill="x", padx=30, pady=10)
        for icon, label, desc in cards: self._make_card(cards_row, icon, label, desc)

    def _make_card(self, parent, icon, label, desc):
        card = ctk.CTkFrame(parent, fg_color="white", corner_radius=12, border_width=1, border_color="#D0E8C0", width=260, height=180); card.pack(side="left", padx=(0, 16), pady=5, fill="both", expand=True); card.pack_propagate(False)
        icon_circle = ctk.CTkFrame(card, fg_color="#C8DDB4", width=60, height=60, corner_radius=30); icon_circle.pack(pady=(20, 10))
        icon_lbl = ctk.CTkLabel(icon_circle, text=icon, font=ctk.CTkFont(size=24)); icon_lbl.place(relx=0.5, rely=0.5, anchor="center")
        title_lbl = ctk.CTkLabel(card, text=label, font=ctk.CTkFont(size=14, weight="bold"), text_color=SECTION_GREEN); title_lbl.pack()
        desc_lbl = ctk.CTkLabel(card, text=desc, font=ctk.CTkFont(size=11), text_color=GRAY_TEXT, wraplength=200, justify="center"); desc_lbl.pack(padx=10)
        
        def trigger(event): self._card_clicked(label)
        card.configure(cursor="hand2"); card.bind("<Button-1>", trigger)
        for w in [icon_circle, icon_lbl, title_lbl, desc_lbl]: w.configure(cursor="hand2"); w.bind("<Button-1>", trigger)
    
    def _view_all_guests(self):
        """
        Opens a new window to browse and search all registered guests.
        """
        win = ctk.CTkToplevel(self)
        win.title("Guest Database Explorer")
        win.geometry("1100x700")
        win.grab_set()  # Focus on this window
        win.configure(fg_color="white")

        # Header
        top_frame = ctk.CTkFrame(win, fg_color="white")
        top_frame.pack(fill="x", padx=30, pady=20)
        ctk.CTkLabel(top_frame, text="Guest Directory", font=ctk.CTkFont(size=22, weight="bold"), text_color=SIDEBAR_GREEN).pack(side="left")
        
        # Search Bar
        search_entry = ctk.CTkEntry(win, placeholder_text="Search by name, email, or username...", width=450)
        search_entry.pack(padx=30, pady=(0, 20), anchor="w")
        
        # Table Header
        header_frame = ctk.CTkFrame(win, fg_color="#E8E8E8", height=40, corner_radius=0)
        header_frame.pack(fill="x", padx=20)
        
        cols = [
            ("Username", 0.02), ("Full Name", 0.15), ("Email", 0.35), 
            ("Phone", 0.55), ("Identity Proof", 0.70), ("Location", 0.85)
        ]
        for text, pos in cols:
            ctk.CTkLabel(header_frame, text=text, font=ctk.CTkFont(size=12, weight="bold")).place(relx=pos, rely=0.5, anchor="w")

        # Scrollable List Container
        list_scroll = ctk.CTkScrollableFrame(win, fg_color="transparent")
        list_scroll.pack(fill="both", expand=True, padx=20, pady=10)

        # Bind search entry to update list in real-time
        search_entry.bind("<KeyRelease>", lambda e: self._update_guest_list(list_scroll, search_entry.get()))
        
        # Initial load of all guests
        self._update_guest_list(list_scroll, "")
    
    def _export_guests_to_excel(self):
        """
        Fetches guest data from the database (respecting active filters) 
        and exports it using the styled excel generator.
        """
        # Get the current search value from the entry widget
        search_val = ""
        if hasattr(self, 'guest_search_entry'):
            search_val = self.guest_search_entry.get().strip()
        
        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()
            
            # SQL Query to match the columns we want in the Excel file
            sql = """SELECT username, first_name, last_name, email, phone, 
                            id_proof_type, id_proof_number, city, country 
                     FROM User WHERE role = 'Guest'"""
            params = []
            
            if search_val:
                sql += " AND (username LIKE %s OR first_name LIKE %s OR last_name LIKE %s OR email LIKE %s)"
                lk = f"%{search_val}%"
                params.extend([lk, lk, lk, lk])
            
            sql += " ORDER BY user_id DESC"
            
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            
            if not rows:
                messagebox.showwarning("Export", "No guest records found to export.")
                return

            # Open File Dialog
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx", 
                filetypes=[("Excel Files", "*.xlsx")], 
                initialfile=f"guest_directory_{datetime.now().strftime('%Y%m%d')}"
            )
            
            if path:
                # Column headers for the Excel sheet
                columns = ["Username", "First Name", "Last Name", "Email", "Phone", 
                           "ID Type", "ID Number", "City", "Country"]
                
                # Convert the database tuples to lists for the generator
                data_rows = [list(r) for r in rows]
                
                # Reusing your professional styled excel logic
                self._create_styled_excel("Guest Directory Report", columns, data_rows, path)
                
                conn.close()
                messagebox.showinfo("Success", f"Guest Directory exported successfully to:\n{os.path.basename(path)}")
                
        except Error as e:
            messagebox.showerror("Database Error", f"Failed to fetch guest data: {e}")


    # --- STAFF REPORTS MODULE (UNIFIED ROOMS & HALLS) ---
    def _open_staff_reports(self):
        win = ctk.CTkToplevel(self); win.title("Staff Performance Reports"); win.geometry("1250x800"); win.grab_set(); win.configure(fg_color="white")
        top_frame = ctk.CTkFrame(win, fg_color="white"); top_frame.pack(fill="x", padx=30, pady=20)
        ctk.CTkLabel(top_frame, text="Staff Processing Report", font=ctk.CTkFont(size=22, weight="bold"), text_color=SIDEBAR_GREEN).pack(side="left")
        ctk.CTkButton(top_frame, text="📥 Export to Excel", fg_color="#2E7D32", width=140, command=self._export_staff_reports_to_excel).pack(side="right", padx=10)
        self.report_search = ctk.CTkEntry(top_frame, placeholder_text="Search staff, guests, or payment...", width=300); self.report_search.pack(side="right", padx=10)
        
        h = ctk.CTkFrame(win, fg_color="#E8E8E8", height=40); h.pack(fill="x", padx=20)
        cols = [
    ("Type", 0.02), 
    ("ID", 0.07), 
    ("Guest Name", 0.14), 
    ("Asset/Room", 0.32), 
    ("Date In", 0.48), 
    ("Processed By", 0.62), 
    ("Status", 0.78),      # Status moved up
    ("Date Paid", 0.89)     # Date Paid is now last
]
        for text, pos in cols: ctk.CTkLabel(h, text=text, font=ctk.CTkFont(size=12, weight="bold")).place(relx=pos, rely=0.5, anchor="w")
        
        report_scroll = ctk.CTkScrollableFrame(win, fg_color="transparent"); report_scroll.pack(fill="both", expand=True, padx=20, pady=10)
        self.report_search.bind("<KeyRelease>", lambda e: self._update_report_list(report_scroll))
        self.current_report_data = []; self._update_report_list(report_scroll)

    def _update_report_list(self, container):
        for w in container.winfo_children(): w.destroy()
        self.current_report_data = []
        q = self.report_search.get().strip()
        try:
            conn = mysql.connector.connect(**DB_CONFIG); cursor = conn.cursor(dictionary=True)
            sql = sql = """
    SELECT 
        'Room' as type, 
        res.reservation_id as id, 
        ug.first_name as gf, 
        ug.last_name as gl, 
        r.room_number as asset, 
        res.check_in_date as d1, 
        us.first_name as sf, 
        us.last_name as sl, 
        res.payment_status,
        res.payment_date  -- Included for ordering
    FROM Reservation res 
    JOIN User ug ON res.user_id = ug.user_id 
    JOIN Room r ON res.room_id = r.room_id
    LEFT JOIN User us ON res.staff_id = us.user_id 
    WHERE r.is_banquet_hall = FALSE

    UNION ALL

    SELECT 
        'Hall' as type, 
        b.booking_id as id, 
        ug.first_name as gf, 
        ug.last_name as gl, 
        r.hall_name as asset, 
        b.event_date as d1, 
        us.first_name as sf, 
        us.last_name as sl, 
        b.payment_status,
        b.payment_date  -- Included for ordering
    FROM Booking b 
    JOIN User ug ON b.user_id = ug.user_id 
    JOIN Room r ON b.room_id = r.room_id
    LEFT JOIN User us ON b.staff_id = us.user_id 
    WHERE r.is_banquet_hall = TRUE

    ORDER BY payment_date DESC
"""
            cursor.execute(sql); rows = cursor.fetchall()
            filtered = [r for r in rows if q.lower() in str(list(r.values())).lower()] if q else rows
            for i, r in enumerate(filtered):
                self.current_report_data.append(r)
                row = ctk.CTkFrame(container, fg_color="white" if i%2==0 else "#F9F9F9", height=45)
                row.pack(fill="x")
                row.pack_propagate(False)

                type_color = "#3498DB" if r['type'] == 'Room' else "#9B59B6"
                ps_color = "#2E7D32" if r['payment_status'] == 'Completed' else "#C62828"

    # Core Data
                ctk.CTkLabel(row, text=r['type'], text_color=type_color, font=ctk.CTkFont(weight="bold")).place(relx=0.02, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=f"#{r['id']}").place(relx=0.07, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=f"{r['gf']} {r['gl']}").place(relx=0.14, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=str(r['asset'])).place(relx=0.32, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=str(r['d1'])).place(relx=0.48, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=f"{r['sf'] or 'Online'}").place(relx=0.62, rely=0.5, anchor="w")

    # Payment Status (Now Penultimate)
                ctk.CTkLabel(row, text=r['payment_status'], text_color=ps_color, font=ctk.CTkFont(weight="bold")).place(relx=0.78, rely=0.5, anchor="w")

    # Payment Date (Now Last)
                ctk.CTkLabel(row, text=str(r['payment_date'] or "N/A")).place(relx=0.89, rely=0.5, anchor="w")
                conn.close()
        except Error as e: print(e)

    def _export_staff_reports_to_excel(self):
        if not self.current_report_data: return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=f"staff_performance_{datetime.now().strftime('%Y%m%d')}")
        if path:
            cols = ["Type", "ID", "Guest First", "Guest Last", "Asset", "Date", "Staff First", "Staff Last", "Payment Status","Date Paid"]
            data_rows = [list(r.values()) for r in self.current_report_data]
            self._create_styled_excel("Staff Performance Report", cols, data_rows, path)
            messagebox.showinfo("Export Success", f"Staff Report saved to {path}")
    def _card_clicked(self, label):
        if label == "Add Guest": self._open_add_guest_form()
        elif label == "Edit Guest": self._open_edit_lookup()
        elif label == "View All Guests": self._view_all_guests()
        elif label == "Add Staff": self._open_add_staff_wizard()
        elif label == "Manage Permissions": self._open_permission_manager()
        elif label == "Staff Reports": self._open_staff_reports() # ADD THIS ELIF BLOCK
        elif label in ["Revenue & Payment Performance", "Occupancy & Utilization Report", "Customer & Booking Behavior"]:
            self._handle_database_export(label)
        else: messagebox.showinfo("Module", f"{label} module coming soon.")

    # --- SHARED PROFESSIONAL EXCEL GENERATOR (FIXED ALIGNMENT & WIDTH) ---
    def _create_styled_excel(self, label, column_names, rows, path):
        wb = Workbook(); ws = wb.active; ws.title = "Report"
        title_fill = PatternFill(start_color="1A0A05", end_color="1A0A05", fill_type="solid")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        alt_row_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        left_align = Alignment(horizontal="left", vertical="center")

        # 1. Dark Title Header
        ws.append([label.upper()])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(column_names))
        ws["A1"].fill = title_fill; ws["A1"].font = Font(color="FFFFFF", bold=True, size=16); ws["A1"].alignment = Alignment(horizontal="center")

        # 2. Metadata (Left Aligned for parity)
        meta = [["Report Name:", label], ["Generated By:", self.admin_name], 
                ["Generated At:", datetime.now().strftime("%B %d, %Y at %I:%M %p")], ["Total Records:", len(rows)]]
        for m_row in meta:
            ws.append(m_row)
            ws[f"A{ws.max_row}"].font = Font(bold=True)
            ws[f"B{ws.max_row}"].alignment = left_align 
        ws.append([]) 

        # 3. Headers
        ws.append(column_names)
        h_idx = ws.max_row
        for cell in ws[h_idx]:
            cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True); cell.border = border; cell.alignment = Alignment(horizontal="center")

        # 4. Data with alternating colors
        for i, r_data in enumerate(rows, start=1):
            ws.append(r_data)
            curr_row = ws.max_row
            for cell in ws[curr_row]:
                cell.border = border
                if i % 2 == 0: cell.fill = alt_row_fill

        # 5. FIXED Autostretch: Skips row 1 merge to avoid error
        for i in range(1, len(column_names) + 1):
            col_letter = get_column_letter(i)
            max_length = 0
            for row in ws.iter_rows(min_row=2, max_col=i, min_col=i):
                for cell in row:
                    if cell.value: max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 8
        wb.save(path)

    def _handle_database_export(self, label):
        mapping = {
            "Revenue & Payment Performance": ("revenue_by_type", """
    SELECT 
        'Room Reservation' AS `Booking/Reservation Type`, 
        r.room_type AS `Booking/Reservation Category`, 
        res.payment_status, 
        res.payment_method, 
        res.payment_date, 
        CONCAT('$', FORMAT(SUM(res.total_amount), 2)) AS revenue
    FROM Room r 
    JOIN Reservation res ON r.room_id = res.room_id 
    WHERE r.is_banquet_hall = FALSE 
      AND res.payment_status = 'COMPLETED' 
    GROUP BY 1, 2, 3, 4, 5

    UNION ALL

    SELECT 
        'Hall Booking' AS `Booking/Reservation Type`, 
        r.room_type AS `Booking/Reservation Category`, 
        b.payment_status, 
        b.payment_method, 
        b.payment_date, 
        CONCAT('$', FORMAT(SUM(b.total_amount), 2)) AS revenue
    FROM Room r 
    JOIN Booking b ON r.room_id = b.room_id 
    WHERE r.is_banquet_hall = TRUE 
      AND b.payment_status = 'COMPLETED' 
    GROUP BY 1, 2, 3, 4, 5

    ORDER BY payment_date DESC, revenue DESC
"""),
            "Occupancy & Utilization Report": ("occupancy_report", """
                SELECT r.room_type, r.room_number, r.status, COUNT(res.reservation_id) AS total_bookings
                FROM Room r LEFT JOIN Reservation res ON r.room_id = res.room_id
                WHERE r.is_banquet_hall = FALSE GROUP BY r.room_id
                UNION ALL
                SELECT  r.hall_name, r.room_number,r.status, COUNT(b.booking_id)
                FROM Room r LEFT JOIN Booking b ON r.room_id = b.room_id
                WHERE r.is_banquet_hall = TRUE GROUP BY r.room_id"""),
           "Customer & Booking Behavior": ("customer_behavior", """
    SELECT 
        u.user_id, 
        u.first_name, 
        u.last_name, 
        u.city, 
        u.country, 
        (
            (SELECT COUNT(*) FROM Reservation res WHERE res.user_id = u.user_id AND res.payment_status = 'Completed') + 
            (SELECT COUNT(*) FROM Booking b WHERE b.user_id = u.user_id AND b.payment_status = 'Completed')
        ) AS `Total Completed Bookings`
    FROM User u 
    WHERE u.role = 'Guest'
    HAVING `Total Completed Bookings` > 0
    ORDER BY `Total Completed Bookings` DESC
""")
        }
        if label not in mapping: return
        file_prefix, query = mapping[label]
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfile=f"{file_prefix}_{datetime.now().strftime('%Y%m%d')}")
        if not path: return
        try:
            conn = mysql.connector.connect(**DB_CONFIG); cursor = conn.cursor()
            cursor.execute(query); rows = cursor.fetchall(); column_names = [i[0].replace("_", " ").title() for i in cursor.description]
            self._create_styled_excel(label, column_names, rows, path)
            conn.close(); messagebox.showinfo("Success", f"{label} Exported.")
        except Exception as e: messagebox.showerror("Error", str(e))

    # PERMISSION MANAGER (GRANT ADMIN ACCESS)
    def _open_permission_manager(self):
        win = ctk.CTkToplevel(self); win.title("Permission Manager"); win.geometry("900x600"); win.grab_set(); win.configure(fg_color="white")
        ctk.CTkLabel(win, text="Manage Staff Access Levels", font=ctk.CTkFont(size=20, weight="bold"), text_color=SIDEBAR_GREEN).pack(pady=20)
        
        search_frame = ctk.CTkFrame(win, fg_color="transparent")
        search_frame.pack(fill="x", padx=40, pady=10)
        
        s_entry = ctk.CTkEntry(search_frame, placeholder_text="Enter Staff Name or Username...", width=400)
        s_entry.pack(side="left", padx=(0, 10))
        
        def run_search(): self._update_staff_permission_list(list_scroll, s_entry.get())
        
        ctk.CTkButton(search_frame, text="Search", fg_color=PRIMARY_GREEN, width=100, command=run_search).pack(side="left")
        
        # Table Header
        h = ctk.CTkFrame(win, fg_color="#E8E8E8", height=40, corner_radius=0); h.pack(fill="x", padx=40, pady=(10, 0))
        cols = [("Name", 0.05), ("Username", 0.30), ("Current Role", 0.55), ("Action", 0.80)]
        for text, pos in cols: ctk.CTkLabel(h, text=text, font=ctk.CTkFont(size=12, weight="bold")).place(relx=pos, rely=0.5, anchor="w")

        list_scroll = ctk.CTkScrollableFrame(win, fg_color="transparent"); list_scroll.pack(fill="both", expand=True, padx=40, pady=10)
        self._update_staff_permission_list(list_scroll, "")

    def _update_staff_permission_list(self, container, query):
        for w in container.winfo_children(): w.destroy()
        try:
            conn = mysql.connector.connect(**DB_CONFIG); cursor = conn.cursor(dictionary=True)
            sql = "SELECT user_id, first_name, last_name, username, role FROM User WHERE role = 'Staff'"
            params = []
            if query.strip():
                sql += " AND (first_name LIKE %s OR last_name LIKE %s OR username LIKE %s)"
                lk = f"%{query}%"
                params.extend([lk, lk, lk])
            
            cursor.execute(sql, params)
            staff_list = cursor.fetchall()
            
            if not staff_list:
                ctk.CTkLabel(container, text="No staff members found.", text_color="gray").pack(pady=20)

            for i, s in enumerate(staff_list):
                row = ctk.CTkFrame(container, fg_color="white" if i%2==0 else "#F9F9F9", height=50, corner_radius=0); row.pack(fill="x"); row.pack_propagate(False)
                ctk.CTkLabel(row, text=f"{s['first_name']} {s['last_name']}").place(relx=0.05, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=s['username']).place(relx=0.30, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=s['role'], text_color="blue").place(relx=0.55, rely=0.5, anchor="w")
                
                # Promotion Button
                ctk.CTkButton(row, text="Grant Admin Access", fg_color="#E67E22", hover_color="#D35400", height=28, width=140, 
                              command=lambda sid=s['user_id'], name=s['username']: self._promote_to_admin(sid, name, container)).place(relx=0.80, rely=0.5, anchor="w")
            conn.close()
        except Error as e: messagebox.showerror("DB Error", str(e))

    def _promote_to_admin(self, user_id, username, container):
        if messagebox.askyesno("Confirm", f"Are you sure you want to grant ADMIN access to {username}?\nThis cannot be undone easily."):
            try:
                conn = mysql.connector.connect(**DB_CONFIG); cursor = conn.cursor()
                cursor.execute("UPDATE User SET role = 'Admin' WHERE user_id = %s", (user_id,))
                conn.commit(); conn.close()
                messagebox.showinfo("Success", f"{username} is now an Administrator.")
                self._update_staff_permission_list(container, "") # Refresh list
            except Error as e: messagebox.showerror("DB Error", str(e))

    # ADD STAFF WIZARD
    def _open_add_staff_wizard(self):
        self.staff_win = ctk.CTkToplevel(self); self.staff_win.title("Staff Recruitment Wizard"); self.staff_win.geometry("550x700"); self.staff_win.grab_set(); self.staff_win.configure(fg_color="white")
        self.staff_data = {}; self.form_container = ctk.CTkFrame(self.staff_win, fg_color="transparent"); self.form_container.pack(fill="both", expand=True)
        self._show_staff_step_1()

    def _show_staff_step_1(self):
        for w in self.form_container.winfo_children(): w.destroy()
        ctk.CTkLabel(self.form_container, text="Step 1: Personal & Account Info", font=ctk.CTkFont(size=18, weight="bold"), text_color=SIDEBAR_GREEN).pack(pady=20)
        scroll = ctk.CTkScrollableFrame(self.form_container, fg_color="transparent"); scroll.pack(fill="both", expand=True, padx=30)
        
        fields = [("Username", "username"), ("Password", "password_hash"), ("Email", "email"), ("First Name", "first_name"), ("Last Name", "last_name"), ("Phone", "phone")]
        self.step1_inputs = {}
        for lbl, col in fields:
            ctk.CTkLabel(scroll, text=lbl, font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(10, 2))
            w = ctk.CTkEntry(scroll, width=400)
            if "password" in col: w.configure(show="*")
            if col in self.staff_data: w.insert(0, self.staff_data[col])
            w.pack(pady=(0, 5)); self.step1_inputs[col] = w
        
        # --- ADDED SECURITY QUESTIONS TO STAFF ---
        ctk.CTkLabel(scroll, text="Security Question 1", font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(15, 2))
        self.q1_var = ctk.StringVar(value=self.staff_data.get("q1", SECURITY_QUESTIONS[0]))
        ctk.CTkOptionMenu(scroll, width=400, values=SECURITY_QUESTIONS, variable=self.q1_var, fg_color=PRIMARY_GREEN).pack(pady=2)
        self.a1_input = ctk.CTkEntry(scroll, width=400, placeholder_text="Answer 1"); self.a1_input.pack(pady=2)
        if "a1" in self.staff_data: self.a1_input.insert(0, self.staff_data["a1"])

        ctk.CTkLabel(scroll, text="Security Question 2", font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(10, 2))
        self.q2_var = ctk.StringVar(value=self.staff_data.get("q2", SECURITY_QUESTIONS[1]))
        ctk.CTkOptionMenu(scroll, width=400, values=SECURITY_QUESTIONS, variable=self.q2_var, fg_color=PRIMARY_GREEN).pack(pady=2)
        self.a2_input = ctk.CTkEntry(scroll, width=400, placeholder_text="Answer 2"); self.a2_input.pack(pady=2)
        if "a2" in self.staff_data: self.a2_input.insert(0, self.staff_data["a2"])

        ctk.CTkButton(self.form_container, text="Next: Employment Details →", fg_color=PRIMARY_GREEN, command=self._proceed_to_staff_step_2).pack(pady=20)

    def _proceed_to_staff_step_2(self):
        # 1. Capture Step 1 Entry fields into the dictionary
        for k, v in self.step1_inputs.items(): 
            self.staff_data[k] = v.get().strip()
        
        # 2. Capture Step 1 Security Questions/Answers
        self.staff_data["q1"] = self.q1_var.get()
        self.staff_data["a1"] = self.a1_input.get().strip()
        self.staff_data["q2"] = self.q2_var.get()
        self.staff_data["a2"] = self.a2_input.get().strip()
        
        # 3. Now move to Step 2 (which clears the UI)
        self._show_staff_step_2()

    def _show_staff_step_2(self):
        for w in self.form_container.winfo_children(): w.destroy()
        ctk.CTkLabel(self.form_container, text="Step 2: Employment Details", font=ctk.CTkFont(size=18, weight="bold"), text_color=SIDEBAR_GREEN).pack(pady=20)
        scroll = ctk.CTkScrollableFrame(self.form_container, fg_color="transparent"); scroll.pack(fill="both", expand=True, padx=30)
        self.step2_vars = {"dept": ctk.StringVar(value=self.staff_data.get("department", "Reception"))}
        ctk.CTkLabel(scroll, text="Department", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        ctk.CTkOptionMenu(scroll, values=["Reception", "Housekeeping", "Events", "Administration", "IT"], variable=self.step2_vars["dept"], fg_color=PRIMARY_GREEN).pack(pady=5)
        self.step2_inputs = {}
        fields = [("Employee ID", "employee_id"), ("Position", "position"), ("Hire Date (YYYY-MM-DD)", "hire_date")]
        for lbl, col in fields:
            ctk.CTkLabel(scroll, text=lbl, font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(10, 2))
            w = ctk.CTkEntry(scroll, width=400); w.pack(pady=(0, 5)); self.step2_inputs[col] = w
            if col == "hire_date" and col not in self.staff_data: w.insert(0, datetime.now().strftime("%Y-%m-%d"))
            elif col in self.staff_data: w.insert(0, self.staff_data[col])
        f = ctk.CTkFrame(self.form_container, fg_color="transparent"); f.pack(pady=20)
        ctk.CTkButton(f, text="← Back", fg_color="gray", width=100, command=self._show_staff_step_1).pack(side="left", padx=10)
        ctk.CTkButton(f, text="Finish & Register", fg_color=PRIMARY_GREEN, width=200, command=self._save_staff).pack(side="left", padx=10)

    def _save_staff(self):
        # Capture remaining fields from Step 2
        self.staff_data["role"] = "Staff"
        self.staff_data["department"] = self.step2_vars["dept"].get()
        for k, v in self.step2_inputs.items(): 
            self.staff_data[k] = v.get().strip()

        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor()
            
            # Precisely 15 columns and 15 placeholders
            sql = """INSERT INTO User (username, password_hash, email, first_name, last_name, 
                                       phone, role, employee_id, position, department, 
                                       hire_date, security_question_1, security_answer_1_hash, 
                                       security_question_2, security_answer_2_hash) 
                     VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            
            val = (
                self.staff_data.get("username"),
                self.staff_data.get("password_hash"),
                self.staff_data.get("email"),
                self.staff_data.get("first_name"),
                self.staff_data.get("last_name"),
                self.staff_data.get("phone"),
                self.staff_data.get("role"),
                self.staff_data.get("employee_id"),
                self.staff_data.get("position"),
                self.staff_data.get("department"),
                self.staff_data.get("hire_date"),
                self.staff_data.get("q1"),
                self.staff_data.get("a1"),
                self.staff_data.get("q2"),
                self.staff_data.get("a2")
            )
            
            cursor.execute(sql, val)
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Success", "Staff member registered successfully.")
            self.staff_win.destroy()
            
        except Error as e:
            messagebox.showerror("Database Error", f"SQL Error: {e}")

    
    # ADD GUEST FORM
    def _open_add_guest_form(self):
        self.form_win = ctk.CTkToplevel(self); self.form_win.title("Guest Registration"); self.form_win.geometry("550x850"); self.form_win.grab_set(); self.form_win.configure(fg_color="white")
        ctk.CTkLabel(self.form_win, text="Create New Guest Account", font=ctk.CTkFont(size=20, weight="bold"), text_color=SIDEBAR_GREEN).pack(pady=(20, 10))
        scroll = ctk.CTkScrollableFrame(self.form_win, fg_color="transparent"); scroll.pack(fill="both", expand=True, padx=25)
        self.inputs = {}
        
        # Base Fields
        fields = [
            ("Username", "username", "entry"), ("Password", "password_hash", "entry"), 
            ("Email", "email", "entry"), ("First Name", "first_name", "entry"), 
            ("Last Name", "last_name", "entry"), ("Phone", "phone", "entry"), 
            ("ID Proof Type", "id_proof_type", "dropdown", ["Passport", "Driver License", "State ID"]), 
            ("ID Proof Number", "id_proof_number", "entry"), ("Address", "address", "entry"), 
            ("City", "city", "entry"), ("Country", "country", "entry")
        ]
        
        for lbl, col, t, *opt in fields:
            ctk.CTkLabel(scroll, text=lbl, font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(10, 2))
            if t == "entry":
                w = ctk.CTkEntry(scroll, width=420); w.pack(pady=(0, 5)); self.inputs[col] = w
                if "password" in col: w.configure(show="*")
            else:
                var = ctk.StringVar(value=opt[0][0]); ctk.CTkOptionMenu(scroll, width=420, values=opt[0], variable=var, fg_color=PRIMARY_GREEN).pack(pady=(0, 5)); self.inputs[col] = var

        # SECURITY QUESTIONS SECTION
        ctk.CTkLabel(scroll, text="Security Recovery Setup", font=ctk.CTkFont(size=14, weight="bold"), text_color=SIDEBAR_GREEN).pack(anchor="w", pady=(20, 5))
        
        # Question 1
        ctk.CTkLabel(scroll, text="Security Question 1", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        self.inputs["q1"] = ctk.StringVar(value=SECURITY_QUESTIONS[0])
        ctk.CTkOptionMenu(scroll, width=420, values=SECURITY_QUESTIONS, variable=self.inputs["q1"], fg_color=PRIMARY_GREEN).pack(pady=5)
        self.inputs["a1"] = ctk.CTkEntry(scroll, placeholder_text="Answer 1", width=420); self.inputs["a1"].pack(pady=(0, 10))

        # Question 2
        ctk.CTkLabel(scroll, text="Security Question 2", font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        self.inputs["q2"] = ctk.StringVar(value=SECURITY_QUESTIONS[1])
        ctk.CTkOptionMenu(scroll, width=420, values=SECURITY_QUESTIONS, variable=self.inputs["q2"], fg_color=PRIMARY_GREEN).pack(pady=5)
        self.inputs["a2"] = ctk.CTkEntry(scroll, placeholder_text="Answer 2", width=420); self.inputs["a2"].pack(pady=(0, 10))

        ctk.CTkButton(scroll, text="Register Guest", fg_color=PRIMARY_GREEN, height=45, width=420, command=self._save_guest).pack(pady=30)

    def _save_guest(self):
        data = {k: (v.get() if isinstance(v, ctk.StringVar) else v.get().strip()) for k, v in self.inputs.items()}
        data["role"] = "Guest"
        
        if not data["a1"] or not data["a2"]:
            messagebox.showerror("Error", "Security answers are required!"); return

        try:
            conn = mysql.connector.connect(**DB_CONFIG); cursor = conn.cursor()
            sql = """INSERT INTO User (username, password_hash, email, first_name, last_name, phone, 
                     id_proof_type, id_proof_number, address, city, country, role, 
                     security_question_1, security_answer_1_hash, security_question_2, security_answer_2_hash) 
                     VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
            val = (data.get("username"), data.get("password_hash"), data.get("email"), data.get("first_name"), 
                   data.get("last_name"), data.get("phone"), data.get("id_proof_type"), data.get("id_proof_number"), 
                   data.get("address"), data.get("city"), data.get("country"), data.get("role"),
                   data.get("q1"), data.get("a1"), data.get("q2"), data.get("a2"))
            cursor.execute(sql, val); conn.commit(); conn.close()
            messagebox.showinfo("Success", "Guest Added."); self.form_win.destroy()
        except Error as e: messagebox.showerror("DB Error", str(e))

    def _update_guest_list(self, container, query_str):
        for widget in container.winfo_children(): widget.destroy()
        try:
            conn = mysql.connector.connect(**DB_CONFIG); cursor = conn.cursor(dictionary=True)
            sql = "SELECT * FROM User WHERE role = 'Guest'"
            params = []
            if query_str.strip():
                sql += " AND (username LIKE %s OR first_name LIKE %s OR last_name LIKE %s OR email LIKE %s)"
                lk = f"%{query_str}%"
                params.extend([lk, lk, lk, lk])
            sql += " ORDER BY user_id DESC"
            cursor.execute(sql, params)
            for i, g in enumerate(cursor.fetchall()):
                row = ctk.CTkFrame(container, fg_color="white" if i%2==0 else "#F9F9F9", height=50, corner_radius=0); row.pack(fill="x"); row.pack_propagate(False)
                ctk.CTkLabel(row, text=g['username']).place(relx=0.02, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=f"{g['first_name']} {g['last_name']}", font=ctk.CTkFont(weight="bold")).place(relx=0.15, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=g['email']).place(relx=0.35, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=g['phone'] or "N/A").place(relx=0.55, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=f"{g['id_proof_type']}: {g['id_proof_number']}").place(relx=0.70, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=f"{g['city']}, {g['country']}").place(relx=0.85, rely=0.5, anchor="w")
            conn.close()
        except Error as e: print(f"Search Error: {e}")

    def _view_all_guests(self):
        """
        Opens a new window to browse and search all registered guests with Export functionality.
        """
        win = ctk.CTkToplevel(self)
        win.title("Guest Database Explorer")
        win.geometry("1150x700")
        win.grab_set()
        win.configure(fg_color="white")

        # Header Frame
        top_frame = ctk.CTkFrame(win, fg_color="white")
        top_frame.pack(fill="x", padx=30, pady=20)
        
        ctk.CTkLabel(top_frame, text="Guest Directory", font=ctk.CTkFont(size=22, weight="bold"), text_color=SIDEBAR_GREEN).pack(side="left")
        
        # --- ADDED EXPORT BUTTON ---
        ctk.CTkButton(top_frame, text="📥 Export Guests", fg_color="#2E7D32", width=140, 
                      command=self._export_guests_to_excel).pack(side="right", padx=10)
        
        # Search Bar
        self.guest_search_entry = ctk.CTkEntry(top_frame, placeholder_text="Search by name, email, or username...", width=350)
        self.guest_search_entry.pack(side="right", padx=10)
        
        # Table Header
        header_frame = ctk.CTkFrame(win, fg_color="#E8E8E8", height=40, corner_radius=0)
        header_frame.pack(fill="x", padx=20)
        
        cols = [
            ("Username", 0.02), ("Full Name", 0.15), ("Email", 0.35), 
            ("Phone", 0.55), ("Identity Proof", 0.70), ("Location", 0.85)
        ]
        for text, pos in cols:
            ctk.CTkLabel(header_frame, text=text, font=ctk.CTkFont(size=12, weight="bold")).place(relx=pos, rely=0.5, anchor="w")

        # Scrollable List Container
        list_scroll = ctk.CTkScrollableFrame(win, fg_color="transparent")
        list_scroll.pack(fill="both", expand=True, padx=20, pady=10)

        # Bind search entry to update list
        self.guest_search_entry.bind("<KeyRelease>", lambda e: self._update_guest_list(list_scroll, self.guest_search_entry.get()))
        
        # Initial load
        self._update_guest_list(list_scroll, "")
    # EDIT LOOKUP
    def _open_edit_lookup(self):
        self.edit_list_win = ctk.CTkToplevel(self)
        self.edit_list_win.title("Select Guest to Edit")
        self.edit_list_win.geometry("900x600")
        self.edit_list_win.grab_set()
        self.edit_list_win.configure(fg_color="white")

        ctk.CTkLabel(self.edit_list_win, text="Guest Editor - Select a Record", 
                     font=ctk.CTkFont(size=20, weight="bold"), text_color=SIDEBAR_GREEN).pack(pady=20)

        # Search bar for narrowing down the list
        search_frame = ctk.CTkFrame(self.edit_list_win, fg_color="transparent")
        search_frame.pack(fill="x", padx=40, pady=10)
        
        edit_search = ctk.CTkEntry(search_frame, placeholder_text="Quick filter by name or email...", width=400)
        edit_search.pack(side="left", padx=(0, 10))

        # Table Header
        h = ctk.CTkFrame(self.edit_list_win, fg_color="#E8E8E8", height=40, corner_radius=0)
        h.pack(fill="x", padx=40, pady=(10, 0))
        cols = [("Full Name", 0.05), ("Username", 0.35), ("Email", 0.55), ("Action", 0.85)]
        for text, pos in cols:
            ctk.CTkLabel(h, text=text, font=ctk.CTkFont(size=12, weight="bold")).place(relx=pos, rely=0.5, anchor="w")

        list_container = ctk.CTkScrollableFrame(self.edit_list_win, fg_color="transparent")
        list_container.pack(fill="both", expand=True, padx=40, pady=10)

        # Bind search
        edit_search.bind("<KeyRelease>", lambda e: self._populate_edit_list(list_container, edit_search.get()))
        
        # Initial load
        self._populate_edit_list(list_container, "")

    def _populate_edit_list(self, container, query):
        for w in container.winfo_children(): w.destroy()
        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cursor = conn.cursor(dictionary=True)
            sql = "SELECT * FROM User WHERE role = 'Guest'"
            params = []
            if query.strip():
                sql += " AND (first_name LIKE %s OR last_name LIKE %s OR username LIKE %s)"
                lk = f"%{query}%"
                params.extend([lk, lk, lk])
            
            cursor.execute(sql, params)
            guests = cursor.fetchall()

            for i, g in enumerate(guests):
                row = ctk.CTkFrame(container, fg_color="white" if i%2==0 else "#F9F9F9", height=50, corner_radius=0)
                row.pack(fill="x")
                row.pack_propagate(False)

                ctk.CTkLabel(row, text=f"{g['first_name']} {g['last_name']}", font=ctk.CTkFont(weight="bold")).place(relx=0.05, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=g['username']).place(relx=0.35, rely=0.5, anchor="w")
                ctk.CTkLabel(row, text=g['email']).place(relx=0.55, rely=0.5, anchor="w")
                
                # The Edit Button (Icon + Text)
                ctk.CTkButton(row, text="✏️ Edit", fg_color=PRIMARY_GREEN, width=80, height=28,
                              command=lambda guest_data=g: self._open_edit_form(guest_data)).place(relx=0.85, rely=0.5, anchor="w")
            conn.close()
        except Error as e: print(e)

    def _fetch_edit(self, u):
        try:
            conn = mysql.connector.connect(**DB_CONFIG); cur = conn.cursor(dictionary=True)
            cur.execute("SELECT * FROM User WHERE username=%s AND role='Guest'", (u,))
            g = cur.fetchone(); conn.close()
            if g: self.win.destroy(); self._open_edit_form(g)
            else: messagebox.showerror("Error", "Not Found.")
        except Error as e: messagebox.showerror("Error", str(e))

    def _open_edit_form(self, g):
        self.e_win = ctk.CTkToplevel(self)
        self.e_win.title(f"Editing: {g['username']}")
        self.e_win.geometry("550x750")
        self.e_win.grab_set()
        
        ctk.CTkLabel(self.e_win, text=f"Update Profile: {g['username']}", 
                     font=ctk.CTkFont(size=18, weight="bold"), text_color=SIDEBAR_GREEN).pack(pady=20)

        scroll = ctk.CTkScrollableFrame(self.e_win, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=25, pady=10)
        
        self.e_inputs = {}
        # Define fields to edit (excluding username for security/integrity)
        fields = [
            ("First Name", "first_name"), ("Last Name", "last_name"), 
            ("Email", "email"), ("Phone", "phone"), 
            ("ID Proof Type", "id_proof_type", ["Passport", "Driver License"]), 
            ("ID Proof Number", "id_proof_number"), ("Address", "address"), 
            ("City", "city"), ("Country", "country")
        ]

        for f in fields:
            ctk.CTkLabel(scroll, text=f[0], font=ctk.CTkFont(weight="bold")).pack(anchor="w", pady=(10, 2))
            if len(f) == 3: # Dropdown type
                var = ctk.StringVar(value=g[f[1]])
                ctk.CTkOptionMenu(scroll, values=f[2], variable=var, fg_color=PRIMARY_GREEN, width=420).pack(pady=5)
                self.e_inputs[f[1]] = var
            else: # Entry type
                w = ctk.CTkEntry(scroll, width=420)
                w.insert(0, str(g[f[1]] or ""))
                w.pack(pady=5)
                self.e_inputs[f[1]] = w

        ctk.CTkButton(scroll, text="Save Changes", fg_color=PRIMARY_GREEN, height=45, width=420, 
                      command=lambda: self._save_edit(g['user_id'])).pack(pady=30)

    def _save_edit(self, uid):
        
        
        d = {k: (v.get() if isinstance(v, ctk.StringVar) else v.get().strip()) for k, v in self.e_inputs.items()}
        try:
            conn = mysql.connector.connect(**DB_CONFIG)
            cur = conn.cursor()
            sql = """UPDATE User SET first_name=%s, last_name=%s, email=%s, phone=%s, 
                     id_proof_type=%s, id_proof_number=%s, address=%s, city=%s, country=%s 
                     WHERE user_id=%s"""
            
            # Ensure the order matches the SQL statement
            values = (d['first_name'], d['last_name'], d['email'], d['phone'], 
                      d['id_proof_type'], d['id_proof_number'], d['address'], 
                      d['city'], d['country'], uid)
            
            cur.execute(sql, values)
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Success", "Guest record updated successfully.")
            self.e_win.destroy()
            # If you want to refresh the selection list behind it:
            if hasattr(self, 'edit_list_win') and self.edit_list_win.winfo_exists():
                self._open_edit_lookup() # This will re-trigger the list view
                
        except Error as e:
            messagebox.showerror("Error", f"Update failed: {e}")
        

    def _sign_out(self):
        path = os.path.join(os.path.dirname(__file__), "session.txt")
        if os.path.exists(path): os.remove(path)
        self.destroy(); subprocess.Popen([sys.executable, os.path.join(os.path.dirname(__file__), "login_screen.py")])
    

if __name__ == "__main__":
    app = UserManagementScreen(); app.mainloop()